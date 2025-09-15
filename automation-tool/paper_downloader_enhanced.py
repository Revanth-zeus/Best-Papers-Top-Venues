import os
import json
import time
import requests
import hashlib
from urllib.parse import urlparse
from collections import defaultdict
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

class PaperDownloader:
    def __init__(self, semantic_scholar_api_key):
        self.ss_api_key = semantic_scholar_api_key
        self.ss_base_url = "https://api.semanticscholar.org/graph/v1"
        self.arxiv_base_url = "http://export.arxiv.org/api/query"
        self.headers = {"x-api-key": self.ss_api_key}
        self.session = requests.Session()
        
        # Statistics tracking
        self.stats = {
            'total_papers': 0,
            'unique_papers': 0,
            'downloaded': 0,
            'already_existed': 0,
            'failed': 0,
            'sections': defaultdict(lambda: {'total': 0, 'downloaded': 0, 'already_existed': 0, 'failed': 0})
        }
        
        # Track unique papers to avoid duplicates
        self.unique_papers = {}  # paper_hash -> paper_info
        self.downloaded_papers = []  # For Excel export
        self.failed_papers = []      # For Excel export
        
        # Create downloads directory
        os.makedirs('downloaded_papers', exist_ok=True)
    
    def clean_title(self, title):
        """Clean title for better matching"""
        return re.sub(r'[^\w\s]', '', title.lower().strip())
    
    def generate_paper_hash(self, title, authors=None):
        """Generate unique hash for paper based on title and authors"""
        clean_title = self.clean_title(title)
        if authors:
            author_str = ''.join(sorted([a.lower().strip() for a in authors if a]))
            hash_input = f"{clean_title}_{author_str}"
        else:
            hash_input = clean_title
        return hashlib.md5(hash_input.encode()).hexdigest()
    
    def fetch_main_readme(self):
        """Fetch main README from GitHub repo"""
        url = "https://raw.githubusercontent.com/SarahRastegar/Best-Papers-Top-Venues/main/README.md"
        try:
            response = requests.get(url)
            if response.status_code == 200:
                return response.text
            else:
                print(f"Failed to fetch main README - Status: {response.status_code}")
                return None
        except Exception as e:
            print(f"Error fetching main README: {e}")
            return None
    
    def is_paper_title_line(self, line):
        """Check if a line contains a paper title (not author info or other metadata)"""
        line = line.strip()
        
        # Skip empty lines
        if not line:
            return False
            
        # Skip obvious non-paper lines
        skip_patterns = [
            r'^authors?:',           # Authors: line
            r'^#',                  # Headers
            r'^best papers?\s',     # "Best Papers" headers
            r'^best paper\s',       # "Best Paper Award" etc
            r'^best student\s',     # "Best Student Paper"
            r'^longuet-higgins\s',  # "Longuet-Higgins Prize"
            r'^test of time\s',     # "Test of Time" headers
            r'^helmholtz\s',        # "Helmholtz Prize"
            r'^outstanding\s',      # "Outstanding Papers"
            r'^award candidates',   # "Award Candidates"
            r'^runners?\s?ups?',    # "Runners up", "Runner-ups"
            r'^contents?$',         # "Contents"
            r'^\[.*\]\[.*\]',       # Link references like [Best Papers] [All Papers]
            r'^back to top',        # Navigation
            r'^table of contents',  # Navigation
            r'^\d{4}\s*$',          # Just years like "2025"
            r'^-+$',                # Separator lines
            r'^=+$',                # Separator lines
            r'^\*+$',               # Separator lines
            r'^\s*\|',              # Table formatting
            r'^venue\s',            # "Venue" headers
            r'^year\s',             # "Year" headers
        ]
        
        for pattern in skip_patterns:
            if re.match(pattern, line, re.IGNORECASE):
                return False
        
        # Must have reasonable length for a paper title
        if len(line) < 10:
            return False
            
        # Should contain some indication it's a paper (has [Paper] or similar patterns)
        # or be a reasonable length title
        has_paper_indicator = bool(re.search(r'\[paper\]|\[pdf\]|\[link\]', line, re.IGNORECASE))
        is_reasonable_title = len(line) > 15 and not line.startswith(('http', 'www', '```'))
        
        return has_paper_indicator or is_reasonable_title
    
    def extract_title_from_line(self, line):
        """Extract clean paper title from a line"""
        line = line.strip()
        
        # Remove markdown link formatting but keep the title
        # Pattern: Title (Conference Year) [Paper]
        match = re.match(r'^(.+?)\s*\([^)]+\)\s*\[.*?\]', line)
        if match:
            title = match.group(1).strip()
        else:
            # Try to extract title from markdown links [Title](URL)
            markdown_match = re.search(r'\[(.+?)\]\([^)]+\)', line)
            if markdown_match:
                title = markdown_match.group(1).strip()
            else:
                # Remove [Paper], [PDF] etc. from end
                title = re.sub(r'\s*\[.*?\]\s*$', '', line).strip()
                # Remove conference and year info in parentheses from end
                title = re.sub(r'\s*\([^)]*\d{4}[^)]*\)\s*$', '', title).strip()
        
        # Clean up any remaining formatting
        title = re.sub(r'\*\*', '', title)  # Remove bold markdown
        title = re.sub(r'\s+', ' ', title)  # Normalize whitespace
        
        return title.strip()
    
    def parse_papers_from_main_readme(self, content):
        """Parse papers from main README content by sections"""
        if not content:
            return {}
        
        sections = {
            'CVPR': [],
            'ICLR': [],
            'NeurIPS': [],
            'ICCV': [],
            'ICML': [],
            'ECCV': [],
            'AAAI': [],
            'WACV': [],
            'BMVC': [],
            'Test-of-Time-Papers': []  # This will capture test of time papers
        }
        
        # Split content into sections
        current_section = None
        lines = content.split('\n')
        
        for line in lines:
            line_stripped = line.strip()
            
            # Check if this line is a section header
            section_found = False
            for section_name in sections.keys():
                # Handle special case for Test of Time Papers
                if section_name == 'Test-of-Time-Papers':
                    # Look for various test of time patterns
                    if re.match(r'#+\s*test\s+of\s+time\s+papers?(?:\s|$)', line_stripped, re.IGNORECASE):
                        current_section = section_name
                        section_found = True
                        print(f"Found section: {section_name}")
                        break
                else:
                    # Regular section matching
                    if re.match(rf'#+\s*{re.escape(section_name)}(?:\s|$)', line_stripped, re.IGNORECASE):
                        current_section = section_name
                        section_found = True
                        print(f"Found section: {section_name}")
                        break
            
            if section_found:
                continue
            
            # If we're in a section and this looks like a paper title line
            if current_section and self.is_paper_title_line(line_stripped):
                title = self.extract_title_from_line(line_stripped)
                
                # Final validation - make sure it's a reasonable title
                if len(title) > 10 and not any(skip in title.lower() for skip in [
                    'readme', 'back to top', 'table of contents', 'best papers', 
                    'best paper', 'authors:', 'venue:', 'year:'
                ]):
                    sections[current_section].append({
                        'title': title,
                        'section': current_section,
                        'raw_line': line
                    })
                    print(f"   Added paper: {title[:60]}...")
        
        return sections
    
    def search_semantic_scholar(self, title):
        """Search for paper on Semantic Scholar"""
        try:
            query_url = f"{self.ss_base_url}/paper/search"
            params = {
                'query': title,
                'limit': 1,
                'fields': 'title,authors,openAccessPdf,externalIds'
            }
            
            response = self.session.get(query_url, headers=self.headers, params=params)
            time.sleep(1)  # Rate limit: 1 request/second
            
            if response.status_code == 200:
                data = response.json()
                if data.get('data') and len(data['data']) > 0:
                    paper = data['data'][0]
                    pdf_url = None
                    
                    if paper.get('openAccessPdf') and paper['openAccessPdf'].get('url'):
                        pdf_url = paper['openAccessPdf']['url']
                    
                    return {
                        'found': True,
                        'pdf_url': pdf_url,
                        'title': paper.get('title', ''),
                        'authors': [a.get('name', '') for a in paper.get('authors', [])],
                        'source': 'semantic_scholar'
                    }
            
            return {'found': False, 'source': 'semantic_scholar'}
            
        except Exception as e:
            print(f"Semantic Scholar search error for '{title[:50]}': {e}")
            return {'found': False, 'error': str(e), 'source': 'semantic_scholar'}
    
    def search_arxiv(self, title):
        """Search for paper on arXiv"""
        try:
            params = {
                'search_query': f'ti:"{title}"',
                'start': 0,
                'max_results': 1
            }
            
            response = requests.get(self.arxiv_base_url, params=params)
            
            if response.status_code == 200:
                content = response.text
                if '<entry>' in content:
                    # Parse basic info from arXiv response
                    pdf_match = re.search(r'<link title="pdf" href="(.*?)"', content)
                    title_match = re.search(r'<title>(.*?)</title>', content)
                    author_matches = re.findall(r'<name>(.*?)</name>', content)
                    
                    if pdf_match:
                        pdf_url = pdf_match.group(1)
                        paper_title = title_match.group(1) if title_match else title
                        authors = author_matches if author_matches else []
                        
                        return {
                            'found': True,
                            'pdf_url': pdf_url,
                            'title': paper_title,
                            'authors': authors,
                            'source': 'arxiv'
                        }
            
            return {'found': False, 'source': 'arxiv'}
            
        except Exception as e:
            print(f"arXiv search error for '{title[:50]}': {e}")
            return {'found': False, 'error': str(e), 'source': 'arxiv'}
    
    def download_pdf(self, pdf_url, filename):
        """Download PDF file with existence check"""
        try:
            filepath = os.path.join('downloaded_papers', filename)
            
            # CHECK IF FILE ALREADY EXISTS
            if os.path.exists(filepath):
                file_size = os.path.getsize(filepath)
                if file_size > 1000:  # Basic check that it's not just an empty/error file
                    print(f"   ⏭️  Already exists: {filename}")
                    return True, "already_exists"
            
            # Continue with download if file doesn't exist
            response = requests.get(pdf_url, stream=True, timeout=30)
            response.raise_for_status()
            
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            
            return True, "downloaded"
            
        except Exception as e:
            print(f"Download error for {filename}: {e}")
            return False, str(e)
    
    def sanitize_filename(self, title, section=''):
        """Create safe filename from paper title"""
        # Remove invalid characters
        safe_title = re.sub(r'[<>:"/\\|?*]', '', title)
        safe_title = safe_title.replace('\n', ' ').replace('\r', ' ')
        safe_title = re.sub(r'\s+', '_', safe_title.strip())
        
        # Limit length
        if len(safe_title) > 80:
            safe_title = safe_title[:80]
        
        # Add section prefix
        if section:
            safe_title = f"{section}_{safe_title}"
        
        return f"{safe_title}.pdf"
    
    def process_papers(self):
        """Main processing function"""
        print("🚀 Starting paper collection and download process...")
        print("=" * 60)
        
        # Step 1: Fetch main README
        print("📄 Fetching main README from GitHub...")
        content = self.fetch_main_readme()
        if not content:
            print("❌ Failed to fetch main README. Exiting.")
            return
        
        print("✅ README fetched successfully!")
        
        # Step 2: Parse papers by sections
        print("\n🔍 Parsing papers from README sections...")
        sections_papers = self.parse_papers_from_main_readme(content)
        
        # Count papers per section
        all_papers = []
        for section, papers in sections_papers.items():
            if papers:
                self.stats['sections'][section]['total'] = len(papers)
                self.stats['total_papers'] += len(papers)
                all_papers.extend(papers)
                print(f"   {section}: {len(papers)} papers")
            else:
                print(f"   {section}: 0 papers")
        
        print(f"\n📊 Total papers found: {self.stats['total_papers']}")
        
        if self.stats['total_papers'] == 0:
            print("❌ No papers found. Please check the README parsing logic.")
            return
        
        # Step 3: Remove duplicates
        print("\n🔄 Removing duplicates...")
        for paper in all_papers:
            paper_hash = self.generate_paper_hash(paper['title'])
            if paper_hash not in self.unique_papers:
                self.unique_papers[paper_hash] = paper
        
        self.stats['unique_papers'] = len(self.unique_papers)
        print(f"   Unique papers after deduplication: {self.stats['unique_papers']}")
        
        # Step 4: Download papers
        print(f"\n⬇️ Starting downloads...")
        print("=" * 60)
        
        for i, (paper_hash, paper) in enumerate(self.unique_papers.items(), 1):
            title = paper['title']
            section = paper['section']
            
            print(f"\n[{i}/{self.stats['unique_papers']}] Processing: {title[:60]}...")
            print(f"   From section: {section}")
            
            # Try Semantic Scholar first
            result = self.search_semantic_scholar(title)
            
            # If not found, try arXiv
            if not result['found']:
                print("   Trying arXiv...")
                result = self.search_arxiv(title)
            
            if result['found'] and result.get('pdf_url'):
                filename = self.sanitize_filename(title, section)
                success, status = self.download_pdf(result['pdf_url'], filename)
                
                if success:
                    if status == "already_exists":
                        print(f"   ⏭️  Already exists: {filename}")
                        self.stats['already_existed'] += 1
                        self.stats['sections'][section]['already_existed'] += 1
                        
                        # Add to downloaded papers list for Excel
                        self.downloaded_papers.append({
                            'Paper Title': title,
                            'Authors': '; '.join(result.get('authors', [])),
                            'Conference/Section': section,
                            'Filename': filename,
                            'Source': result.get('source', ''),
                            'Status': 'Already Existed'
                        })
                    else:
                        print(f"   ✅ Downloaded: {filename}")
                        self.stats['downloaded'] += 1
                        self.stats['sections'][section]['downloaded'] += 1
                        
                        # Add to downloaded papers list for Excel
                        self.downloaded_papers.append({
                            'Paper Title': title,
                            'Authors': '; '.join(result.get('authors', [])),
                            'Conference/Section': section,
                            'Filename': filename,
                            'Source': result.get('source', ''),
                            'Status': 'Downloaded'
                        })
                else:
                    print(f"   ❌ Download failed: {status}")
                    self.stats['failed'] += 1
                    self.stats['sections'][section]['failed'] += 1
                    
                    # Add to failed papers list for Excel
                    self.failed_papers.append({
                        'Paper Title': title,
                        'Authors': '; '.join(result.get('authors', [])),
                        'Conference/Section': section,
                        'Reason': f'Download failed: {status}',
                        'PDF URL': result.get('pdf_url', ''),
                        'Source': result.get('source', '')
                    })
            else:
                print(f"   ❌ PDF not found")
                self.stats['failed'] += 1
                self.stats['sections'][section]['failed'] += 1
                
                # Add to failed papers list for Excel
                self.failed_papers.append({
                    'Paper Title': title,
                    'Authors': '; '.join(result.get('authors', [])) if result.get('authors') else '',
                    'Conference/Section': section,
                    'Reason': 'PDF not found',
                    'PDF URL': '',
                    'Source': result.get('source', '')
                })
        
        # Step 5: Generate reports
        self.generate_reports()
        self.create_excel_report()
    
    def generate_reports(self):
        """Generate final reports and statistics"""
        print("\n" + "=" * 60)
        print("📋 FINAL STATISTICS")
        print("=" * 60)
        
        print(f"Total papers found: {self.stats['total_papers']}")
        print(f"Unique papers: {self.stats['unique_papers']}")
        print(f"Successfully downloaded: {self.stats['downloaded']}")
        print(f"Already existed: {self.stats['already_existed']}")
        print(f"Failed downloads: {self.stats['failed']}")
        
        if self.stats['unique_papers'] > 0:
            total_success = self.stats['downloaded'] + self.stats['already_existed']
            print(f"Overall success rate: {(total_success/self.stats['unique_papers']*100):.1f}%")
        
        print(f"\n📁 PER SECTION STATISTICS:")
        print("-" * 50)
        for section, stats in self.stats['sections'].items():
            if stats['total'] > 0:
                print(f"{section}:")
                print(f"   Total: {stats['total']}")
                print(f"   Downloaded: {stats['downloaded']}")
                print(f"   Already existed: {stats['already_existed']}")
                print(f"   Failed: {stats['failed']}")
                total_success = stats['downloaded'] + stats['already_existed']
                success_rate = (total_success / stats['total'] * 100) if stats['total'] > 0 else 0
                print(f"   Success rate: {success_rate:.1f}%")
                print()
        
        # Save detailed results to JSON
        report = {
            'summary': dict(self.stats),  # Convert defaultdict to regular dict
            'downloaded_count': len(self.downloaded_papers),
            'failed_count': len(self.failed_papers),
            'timestamp': time.strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Convert defaultdict to regular dict for JSON serialization
        report['summary']['sections'] = dict(self.stats['sections'])
        
        with open('download_report.json', 'w') as f:
            json.dump(report, f, indent=2)
        
        print(f"📄 Detailed report saved to: download_report.json")
        print(f"📁 Downloaded papers saved to: downloaded_papers/")
    
    def create_excel_report(self):
        """Create Excel file with two sheets: Downloaded and Failed papers"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create Downloaded Papers sheet
            ws_downloaded = wb.create_sheet(title="Downloaded Papers")
            
            if self.downloaded_papers:
                # Convert to DataFrame for easier handling
                df_downloaded = pd.DataFrame(self.downloaded_papers)
                
                # Add headers with styling
                headers = list(df_downloaded.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws_downloaded.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for row_num, row_data in enumerate(df_downloaded.itertuples(index=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        ws_downloaded.cell(row=row_num, column=col_num, value=value)
                
                # Auto-adjust column widths
                for column in ws_downloaded.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws_downloaded.column_dimensions[column_letter].width = adjusted_width
            else:
                ws_downloaded.cell(row=1, column=1, value="No papers were downloaded")
            
            # Create Failed Papers sheet
            ws_failed = wb.create_sheet(title="Failed Papers")
            
            if self.failed_papers:
                # Convert to DataFrame for easier handling
                df_failed = pd.DataFrame(self.failed_papers)
                
                # Add headers with styling
                headers = list(df_failed.columns)
                for col_num, header in enumerate(headers, 1):
                    cell = ws_failed.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Add data
                for row_num, row_data in enumerate(df_failed.itertuples(index=False), 2):
                    for col_num, value in enumerate(row_data, 1):
                        ws_failed.cell(row=row_num, column=col_num, value=value)
                
                # Auto-adjust column widths
                for column in ws_failed.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws_failed.column_dimensions[column_letter].width = adjusted_width
            else:
                ws_failed.cell(row=1, column=1, value="No failed downloads")
            
            # Save Excel file
            excel_filename = f"paper_download_results_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(excel_filename)
            
            print(f"📊 Excel report saved to: {excel_filename}")
            print(f"   📑 Sheet 1: Downloaded Papers ({len(self.downloaded_papers)} papers)")
            print(f"   📑 Sheet 2: Failed Papers ({len(self.failed_papers)} papers)")
            
        except Exception as e:
            print(f"❌ Error creating Excel report: {e}")
            print("📄 Falling back to CSV files...")
            
            # Create CSV files as fallback
            if self.downloaded_papers:
                df_downloaded = pd.DataFrame(self.downloaded_papers)
                csv_downloaded = f"downloaded_papers_{time.strftime('%Y%m%d_%H%M%S')}.csv"
                df_downloaded.to_csv(csv_downloaded, index=False)
                print(f"📊 Downloaded papers CSV saved to: {csv_downloaded}")
            
            if self.failed_papers:
                df_failed = pd.DataFrame(self.failed_papers)
                csv_failed = f"failed_papers_{time.strftime('%Y%m%d_%H%M%S')}.csv"
                df_failed.to_csv(csv_failed, index=False)
                print(f"📊 Failed papers CSV saved to: {csv_failed}")


# Main execution
if __name__ == "__main__":
    # Check if required packages are installed
    try:
        import pandas as pd
        import openpyxl
    except ImportError as e:
        print(f"❌ Missing required package: {e}")
        print("Please install required packages:")
        print("pip install pandas openpyxl")
        exit(1)
    
    API_KEY = "dFjbDAiB6h3fT8UnUDqv31HaWQY8zJHw6sKpVSth"
    
    downloader = PaperDownloader(API_KEY)
    downloader.process_papers()
    
    print("\n🎉 Process completed!")
    print("Check the 'downloaded_papers' folder for your PDFs.")
    print("Check the Excel file for organized results with paper details! 📊")